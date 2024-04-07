import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
import os
import openpyxl
from rdkit import Chem
from rdkit.Chem import Draw


from props_ptr.name_chem import chem_name_info
from props_ptr.mol_prop import basic_mol_props
from props_ptr.mol_struct import calculate_molecular_parameters
from props_ptr.mol_exp import exp_info
from props_ptr.graphs_1 import graph_prop_1,plot_radar_chart_web
from props_ptr.database_link import get_database_links
from props_ptr.sa_dl import calculate_sa_score
from props_ptr.struct_mdc import struct_exp
from props_ptr.drug_fil import rules_drg
from props_ptr.ESOL_prop import predict_sol
from props_ptr.pharmadmet import predict_for_all_models

def param_src(smiles):
    molecule = Chem.MolFromSmiles(smiles)
    params_name = chem_name_info(smiles)
    mol_prop_func = basic_mol_props(smiles)
    mol_struct_func = calculate_molecular_parameters(smiles)
    sa_func = calculate_sa_score(smiles)
    drg_struc_func = struct_exp(smiles)
    drg_lik_func = rules_drg(smiles)
    phys_func = predict_sol(smiles)
    exp_name = exp_info(smiles)
    database_links = get_database_links(smiles)
    compound_values = graph_prop_1(smiles)
    parma = predict_for_all_models(smiles)
    canvas_plot = plot_radar_chart_web(compound_values)
    img_chem = Draw.MolToImage(molecule, size=(300, 200))
    img_chem.save("temp_img/molecule.png")
    molecule_data = {
    "Names and Identifiers": params_name,
    "Molecular Properties": mol_prop_func,
    "Structural Properties": mol_struct_func,
    "Accessibility Score": sa_func,
    "Drug-like Structural Features": drg_struc_func,
    "Drug-likeness Rules": drg_lik_func,
    "Physicochemical Properties": phys_func,
    "Pharmacokinetics Properties": parma,
    "Record Properties": exp_name,
    "Database Links": database_links
}
    return molecule_data


def adjust_for_excel(data):
    excel_ready_data = {}
    for section, content in data.items():
        if content is not None:
            excel_ready_data[section] = [dict(zip(["Property", "Value"], item)) for item in content.items()]
        else:
            excel_ready_data[section] = [{"Property": "N/A", "Value": "N/A"}]
    return excel_ready_data


def adjust_for_pdf(data):
    return data  # No adjustment needed for PDF export

def export_to_excel(data, filename="export.xlsx"):
    adjusted_data = adjust_for_excel(data)
    with pd.ExcelWriter(filename) as writer:
        for section, rows in adjusted_data.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=section, index=False)

def export_to_excel_with_image(data, excel_filename, molecule_image_filename='temp_img/molecule.png', radar_image_filename='temp_img/radar_chart.png'):
    # Export data to Excel using pandas
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
    for section, rows in data.items():
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name=section, index=False)
    writer.close() 

    # Open the Excel file with openpyxl for post-processing
    wb = openpyxl.load_workbook(excel_filename)

    # Assuming the first sheet and the fifth sheet exist
    ws_first = wb[wb.sheetnames[0]]
    img_molecule = openpyxl.drawing.image.Image(molecule_image_filename)
    img_molecule.anchor = 'A11'  # Adjust as needed
    ws_first.add_image(img_molecule)

    ws_fifth = wb[wb.sheetnames[5]]  # Correct index for 5th sheet
    img_radar = openpyxl.drawing.image.Image(radar_image_filename)
    img_radar.anchor = 'A11'  # Adjust as needed
    ws_fifth.add_image(img_radar)

    # Save the workbook
    wb.save(excel_filename)


def export_to_pdf(data, filename="export.pdf", molecule_image_filename='temp_img/molecule.png', radar_image_filename='temp_img/radar_chart.png'):
    adjusted_data = adjust_for_pdf(data)
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()

    # Molecule image on the first page
    story = []
    story.append(Paragraph("Molecule Result Data", styles['Title']))
    molecule_img = Image(molecule_image_filename, width=200, height=150)
    story.append(molecule_img)
    story.append(Spacer(1, 12))

    # Adding data
    for section, content in adjusted_data.items():
        story.append(Paragraph(section, styles['Heading2']))
        if content is not None:
            for key, value in content.items():
                story.append(Paragraph(f"{key}: {value}", styles['Normal']))
                story.append(Spacer(1, 12))

    # Spiderweb image on the last page
    radar_img = Image(radar_image_filename, width=500, height=300)
    story.append(Spacer(1, 12))
    story.append(radar_img)

    doc.build(story)

def export_mol_pdf(smiles, filename):
    molecule_data = param_src(smiles)
    # Use the provided filename for PDF export
    export_to_pdf(molecule_data, filename)

def export_mol_xls(smiles, filename):
    molecule_data = param_src(smiles)
    adjusted_data = adjust_for_excel(molecule_data)
    # Use the provided filename for Excel export, which includes dynamically adding images
    export_to_excel_with_image(adjusted_data, filename, "temp_img/molecule.png", "temp_img/radar_chart.png")

